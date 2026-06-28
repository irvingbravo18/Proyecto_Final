import random
import string
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime


#VALIDAMOS RESPUESTAS DEL USUARIO

def Validar_Respuesta(msg):
    while True:
        r=input(msg).lower()
        if r in ("s","n"):
            return r=="s"
        print("")
        print("=======================================================================")
        print("ESTIMADO USUARIO SE GENERO UN ERROR USTED SOLO PUEDE INGRESAR: S o N.")
        print("=======================================================================")
        print("")

def generar(Longitud,Mayuscula, Minuscula, Numero, Simbolo):
    chars=""
    if Mayuscula: chars+=string.ascii_uppercase
    if Minuscula: chars+=string.ascii_lowercase
    if Numero: chars+=string.digits
    if Simbolo: chars+=string.punctuation
    if not chars: return None
    return "".join(random.choice(chars) for _ in range(Longitud))

def Evaluar(p):
    n=0
    if len(p)>=8:n+=1
    if any(c.isupper() for c in p): n+=1
    if any(c.islower() for c in p): n+=1
    if any(c.isdigit() for c in p): n+=1
    if any(c in string.punctuation for c in p): n+=1
    return "NIVEL DE SEGURIDAD ALTO" if n==5 else "NIVEL DE SEGURIDAD MEDIO" if n>=3 else "NIVEL DE SEGURIDAD BAJO"

ruta=os.path.join(os.path.expanduser("~"),"Desktop","Historial.xlsx")
if not os.path.exists(ruta):
    wb=Workbook(); ws=wb.active
    ws.append(["Contraseña","Nivel","Fecha","Hora"])
    wb.save(ruta)

while True:
    print("")
    print("==============================================")
    print("     GENERADOR DE CONTRASEÑAS")
    print("==============================================")
    print("")
    print("SE PRESENTA EL SIGUINETE MENÚ")
    print("")
    print("1. Generar Contraseña")
    print("2. Salir")
    print("")
    op=input("SELECIONE UNA DE LAS SIGUINETES OPCIONES: ")
    print("")
    if op=="1":
        try:
            print("")
            l=int(input("CUANTOS CARACTERES DESEA QUE CONTENGA SU CONTRASEÑA (8-20): "))
            print("")
            if l<8 or l>20:
                    print("=======================================================================")
                    print("Debe ingresar un numero que defina una  dimencion correcta"); continue
                    
            print("")
                    
        except:
            print("Ingrese un número"); continue
        print("")
        print("=======================================================================")
        print("ESCRIBA EN EL TECLADO *** S *** SI USTED DESEA AGREGAR")
        print("ESCRIBA EN EL TECLADO *** N *** SI NO DESEA AGREGAR")
        print("=======================================================================")
        print("")
        Mayuscula=Validar_Respuesta("¿DESEA QUE SU CONTRASEÑA CONTENGA LETRAS MAYUSCULAS? (s/n): ")
        print("")
        Minuscula=Validar_Respuesta("¿DESEA QUE SU CONTRASEÑA CONTENGA LETRAS MINUSCULA?  (s/n): ")
        print("")
        Numero=Validar_Respuesta("¿DESEA QUE SU CONTRASEÑA CONTENGA NÚMEROS?           (s/n): ")
        print("")
        Simbolo=Validar_Respuesta("¿DESEA QUE SU CONTRASEÑA CONTENGA SIMBOLOS?          (s/n): ")
        print("")
        pwd=generar(l, Mayuscula, Minuscula, Numero, Simbolo)
        if pwd is None:
            print("SELECIONE UNA DE LAS SIGUIENTES OPCIONES."); 
        print("")
        nivel=Evaluar(pwd)
        print("") 
        print("Contraseña:",pwd)
        print("")
        print("Nivel:",nivel)
        print("")
        print("=====================================================")
        if Validar_Respuesta("¿DESEA GUARDAR LA INFORMACION? (s/n): "):
            print("=================================================")
            print("")
            wb=load_workbook(ruta); ws=wb.active
            now=datetime.now()
            ws.append([pwd,nivel,now.strftime("%d/%m/%Y"),now.strftime("%H:%M:%S")])
            wb.save(ruta)
            print("")
            print("Guardada en",ruta)
            print("")
    elif op=="2":
        break
    else:
        print("========================================================================================")
        print("Estimado Usuario selecione el numero correcto de acuerdo al menú mostrado a continuacion")
        print("========================================================================================")
